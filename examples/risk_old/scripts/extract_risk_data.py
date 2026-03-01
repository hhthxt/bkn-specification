#!/usr/bin/env python3
"""
Extract risk data from case files (PFMEA, DFMEA, 风险表, docx) into BKN-aligned CSV files.
Output: scenario.csv, action_option.csv, risk.csv, risk_statement.csv, and 3 relation CSVs.
"""

import csv
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from docx import Document
except ImportError:
    Document = None


def sanitize_id(s: str) -> str:
    """Convert string to valid BKN ID: lowercase, digits, underscore, hyphen."""
    if not s:
        return ""
    s = re.sub(r"[^\w\-]", "_", s.lower().strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:64] if s else ""


def map_severity_to_1_5(val) -> int | None:
    """Map 1-9/1-10 scale to 1-5."""
    try:
        v = int(float(val))
        if v <= 0:
            return None
        if v <= 2:
            return 1
        if v <= 4:
            return 2
        if v <= 6:
            return 3
        if v <= 8:
            return 4
        return 5
    except (ValueError, TypeError):
        return None


def map_risk_level(val: str) -> int | None:
    """Map 极高/高/中/低 to 1-5."""
    m = {"极高": 5, "高": 4, "中": 3, "低": 2}
    return m.get((val or "").strip(), None)


def infer_category(text: str) -> str:
    """Infer BKN category from text."""
    t = (text or "").lower()
    if any(x in t for x in ["安全", "泄露", "授权", "权限"]):
        return "security"
    if any(x in t for x in ["数据丢失", "数据覆盖", "无法恢复", "回滚"]):
        return "integrity"
    if any(x in t for x in ["可用", "中断", "宕机", "不可用", "失败"]):
        return "availability"
    if any(x in t for x in ["性能", "慢", "卡顿", "瓶颈"]):
        return "performance"
    if any(x in t for x in ["依赖", "网络", "连接"]):
        return "dependency"
    return "operator"


def infer_risk_type(text: str) -> str:
    """Map text to risk_type enum."""
    t = (text or "").lower()
    if any(x in t for x in ["数据丢失", "无法恢复", "销毁", "删除"]):
        return "data_loss"
    if any(x in t for x in ["不一致", "覆盖", "回滚"]):
        return "inconsistency"
    if any(x in t for x in ["可用", "中断", "宕机", "不可用"]):
        return "availability"
    if any(x in t for x in ["安全", "泄露", "授权"]):
        return "security"
    if any(x in t for x in ["合规"]):
        return "compliance"
    if any(x in t for x in ["财务"]):
        return "financial"
    if any(x in t for x in ["声誉"]):
        return "reputation"
    return "availability"


def infer_action_type(text: str) -> str:
    """Map text to action_type enum."""
    t = (text or "").lower()
    if any(x in t for x in ["failover", "故障转移", "切换"]):
        return "failover"
    if any(x in t for x in ["wait", "等待", "观察"]):
        return "wait"
    if any(x in t for x in ["restore", "恢复", "还原"]):
        return "restore"
    if any(x in t for x in ["rollback", "回滚"]):
        return "rollback"
    if any(x in t for x in ["degrade", "降级"]):
        return "degrade"
    if any(x in t for x in ["isolate", "隔离"]):
        return "isolate"
    if any(x in t for x in ["rebuild", "重建"]):
        return "rebuild"
    return "restore"


def infer_reversibility(text: str) -> str:
    """Infer reversibility from risk level or action."""
    t = (text or "").lower()
    if any(x in t for x in ["删除", "格式化", "truncate", "prune", "rm -rf"]):
        return "irreversible"
    if any(x in t for x in ["修改", "chmod", "chown"]):
        return "partially_reversible"
    return "reversible"


# --- Data collectors ---

scenarios: dict[str, dict] = {}
action_options: dict[str, dict] = {}
risks: dict[str, dict] = {}
risk_statements: list[dict] = []


def ensure_scenario(scenario_id: str, name: str, category: str, primary_object: str, description: str = ""):
    if scenario_id and scenario_id not in scenarios:
        scenarios[scenario_id] = {
            "scenario_id": scenario_id,
            "name": name or scenario_id,
            "category": category or "operator",
            "primary_object": primary_object or "system",
            "description": description or "",
        }


def ensure_action(action_id: str, name: str, action_type: str, reversibility: str, runbook_ref: str = "", description: str = ""):
    if action_id and action_id not in action_options:
        action_options[action_id] = {
            "action_id": action_id,
            "name": name or action_id,
            "action_type": action_type or "restore",
            "reversibility": reversibility or "partially_reversible",
            "runbook_ref": runbook_ref or "",
            "description": description or "",
        }


def ensure_risk(risk_id: str, name: str, risk_type: str, description: str = ""):
    if risk_id and risk_id not in risks:
        risks[risk_id] = {
            "risk_id": risk_id,
            "name": name or risk_id,
            "risk_type": risk_type or "availability",
            "description": description or "",
        }


def add_risk_statement(
    rs_id: str,
    name: str,
    scenario_id: str,
    action_id: str,
    risk_id: str,
    likelihood_level: int | None = None,
    business_impact: int | None = None,
    data_impact: int | None = None,
    compliance_impact: int | None = None,
    notes: str = "",
    mitigation_pre: str = "",
    mitigation_pro: str = "",
):
    risk_statements.append({
        "rs_id": rs_id,
        "name": name[:200] if name else rs_id,
        "status": "active",
        "scenario_id": scenario_id,
        "action_id": action_id,
        "risk_id": risk_id,
        "likelihood_level": likelihood_level or "",
        "business_impact": business_impact or "",
        "data_impact": data_impact or "",
        "compliance_impact": compliance_impact or "",
        "notes": (notes or "")[:500],
        "mitigation_pre": (mitigation_pre or "")[:500],
        "mitigation_pro": (mitigation_pro or "")[:500],
    })


# --- Extract from 风险表.xlsx ---

def extract_risk_table(path: Path):
    if not openpyxl:
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in ["程序运行类", "系统运维类"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else []
        for i, row in enumerate(rows[1:], start=2):
            vals = [str(c) if c is not None else "" for c in (row or [])]
            if len(vals) < 4:
                continue
            # 程序运行类: 风险等级, 命令/操作范围, 适用版本, 允许/禁止, 风险说明
            # 系统运维类: 类别, 风险等级, 命令/操作范围, 适用版本, 允许/禁止, 说明
            if sheet_name == "程序运行类":
                risk_lvl, cmd, ver, allow, desc = vals[0], vals[1], vals[2], vals[3], vals[4] if len(vals) > 4 else ""
                category_raw = "程序运行"
            else:
                cat_raw, risk_lvl, cmd, ver, allow, desc = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5] if len(vals) > 5 else ""
                category_raw = cat_raw or "系统运维"
            if not cmd.strip():
                continue
            sid = sanitize_id(f"ops-{category_raw}-{i}")
            aid = sanitize_id(f"ops-cmd-{i}")
            rid = sanitize_id(f"ops-risk-{infer_risk_type(desc)}-{i}")
            rsid = f"ops-{i}"
            ensure_scenario(sid, f"{category_raw}-{cmd[:30]}", infer_category(desc), "backup_system", desc)
            ensure_action(aid, cmd[:80], "restore", infer_reversibility(cmd), "", desc)
            ensure_risk(rid, desc[:80] if desc else "操作风险", infer_risk_type(desc), desc)
            add_risk_statement(
                rsid, f"{category_raw}: {cmd[:50]}",
                sid, aid, rid,
                business_impact=map_risk_level(risk_lvl),
                notes=f"适用版本:{ver} 允许/禁止:{allow}",
                mitigation_pre=desc,
            )


# --- Extract from PFMEA ---

def extract_pfmea(path: Path):
    if not openpyxl:
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    # Skip 实时&容灾 - has 1M+ rows, only ~1 data row; add that row manually below
    data_sheets = ["定时备份", "云备份", "磁带备份", "公共管理类", "其他"]
    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))[:150]
        for i, row in enumerate(rows[1:], start=2):
            vals = [str(c) if c is not None else "" for c in (row or [])]
            if len(vals) < 8:
                continue
            mod, check, fail_mode, link, ctrl, sev, freq, detect = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], vals[6], vals[7]
            improve = vals[9] if len(vals) > 9 else ""
            if not check.strip() and not fail_mode.strip():
                continue
            mod = mod or sheet_name
            sid = sanitize_id(f"pfmea-{mod}-{i}")
            aid = sanitize_id(f"pfmea-action-{mod}-{i}")
            rid = sanitize_id(f"pfmea-risk-{fail_mode}-{i}")
            rsid = f"pfmea-{sanitize_id(mod)}-{i}"
            ensure_scenario(sid, f"{mod}: {check[:40]}", infer_category(check + fail_mode), mod, check)
            ensure_action(aid, improve[:80] if improve else f"应对{fail_mode}", infer_action_type(improve or fail_mode), "partially_reversible", "", improve)
            ensure_risk(rid, fail_mode[:80] if fail_mode else "失效", infer_risk_type(fail_mode + check), check)
            add_risk_statement(
                rsid, f"{mod}: {fail_mode[:50]}",
                sid, aid, rid,
                likelihood_level=map_severity_to_1_5(freq),
                business_impact=map_severity_to_1_5(sev),
                data_impact=map_severity_to_1_5(sev),
                notes=f"RPN; 现有控制:{ctrl}" if ctrl else "",
                mitigation_pre=ctrl,
                mitigation_pro=improve,
            )

    # 实时&容灾 sheet has 1M+ rows, only 1 data row - add manually
    ensure_scenario("pfmea-realtime-1", "实时: 恢复覆盖风险", "integrity", "实时", "将备份数据恢复到重名的文件或目录下会覆盖数据")
    ensure_action("pfmea-realtime-action-1", "添加最佳实践", "restore", "partially_reversible", "", "恢复前检查目标路径")
    ensure_risk("pfmea-realtime-risk-1", "数据覆盖", "inconsistency", "恢复时覆盖现有数据")
    add_risk_statement(
        "pfmea-realtime-1", "实时: 数据覆盖",
        "pfmea-realtime-1", "pfmea-realtime-action-1", "pfmea-realtime-risk-1",
        likelihood_level=2, business_impact=5, data_impact=5,
        notes="RPN 162; 产品设计",
        mitigation_pre="添加最佳实践",
    )


# --- Extract from DFMEA ---

def extract_dfmea(path: Path):
    if not openpyxl:
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    data_sheets = ["KVM数据备份评审", "数据恢复评审", "容灾接管评审"]
    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[1:], start=2):
            vals = [str(c) if c is not None else "" for c in (row or [])]
            if len(vals) < 8:
                continue
            func, fail_mode, impact = vals[0], vals[1], vals[2]
            sev, cause, freq, ctrl, detect = vals[3], vals[4], vals[5], vals[6], vals[7]
            improve = vals[8] if len(vals) > 8 else ""
            if not fail_mode.strip() and not impact.strip():
                continue
            func = func or sheet_name
            sid = sanitize_id(f"dfmea-{func}-{i}")
            aid = sanitize_id(f"dfmea-action-{func}-{i}")
            rid = sanitize_id(f"dfmea-risk-{fail_mode}-{i}")
            rsid = f"dfmea-{sanitize_id(func)}-{i}"
            ensure_scenario(sid, f"{func}: {fail_mode[:40]}", infer_category(impact + fail_mode), func, impact)
            ensure_action(aid, improve[:80] if improve else f"应对{fail_mode}", infer_action_type(improve or fail_mode), "partially_reversible", "", improve)
            ensure_risk(rid, fail_mode[:80] if fail_mode else "失效", infer_risk_type(impact + fail_mode), impact)
            add_risk_statement(
                rsid, f"{func}: {fail_mode[:50]}",
                sid, aid, rid,
                likelihood_level=map_severity_to_1_5(freq),
                business_impact=map_severity_to_1_5(sev),
                data_impact=map_severity_to_1_5(sev),
                notes=f"起因:{cause} 控制:{ctrl}" if cause or ctrl else "",
                mitigation_pre=ctrl,
                mitigation_pro=improve,
            )


# --- Extract from docx ---

def extract_docx(path: Path):
    if not Document:
        return
    doc = Document(path)
    current_section = ""
    idx = [0]  # mutable to allow closure

    def next_id():
        idx[0] += 1
        return idx[0]

    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if not t or t == "0000":
            continue
        if p.style.name and "Heading" in p.style.name:
            current_section = t
            continue
        i = next_id()
        sec = sanitize_id(current_section) or "decl"
        sid = f"decl-{sec}-{i}"
        aid = f"decl-action-{i}"
        rid = f"decl-risk-{infer_risk_type(t)}-{i}"
        rsid = f"decl-{sec}-{i}"
        ensure_scenario(sid, current_section or "风险声明", infer_category(t), current_section or "product", t)
        ensure_action(aid, "遵循产品说明操作", "wait", "reversible", "", t[:200])
        ensure_risk(rid, t[:80], infer_risk_type(t), t)
        add_risk_statement(rsid, t[:80], sid, aid, rid, notes=t[:300], mitigation_pre=t)


# --- Main ---

def main():
    base = Path(__file__).resolve().parent.parent
    cases_dir = base / "cases"
    data_dir = base / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    # Reset
    scenarios.clear()
    action_options.clear()
    risks.clear()
    risk_statements.clear()

    # Extract
    risk_table = cases_dir / "风险表.xlsx"
    pfmea = cases_dir / "PFMEA模板.xlsx"
    dfmea = cases_dir / "DFMEA模板.xlsx"
    docx_path = cases_dir / "AnyBackup 7.0.1.0 风险声明文档.docx"

    if risk_table.exists():
        extract_risk_table(risk_table)
    if pfmea.exists():
        extract_pfmea(pfmea)
    if dfmea.exists():
        extract_dfmea(dfmea)
    if docx_path.exists():
        extract_docx(docx_path)

    # Deduplicate risk_statement rs_id (avoid collisions)
    seen_rs = set()
    unique_rs = []
    for r in risk_statements:
        rid = r["rs_id"]
        if rid in seen_rs:
            rid = f"{rid}-{len(seen_rs)}"
            r["rs_id"] = rid
        seen_rs.add(r["rs_id"])
        unique_rs.append(r)

    # Write CSVs
    def write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    write_csv(data_dir / "scenario.csv", list(scenarios.values()),
              ["scenario_id", "name", "category", "primary_object", "description"])
    write_csv(data_dir / "action_option.csv", list(action_options.values()),
              ["action_id", "name", "action_type", "reversibility", "runbook_ref", "description"])
    write_csv(data_dir / "risk.csv", list(risks.values()),
              ["risk_id", "name", "risk_type", "description"])

    rs_fields = ["rs_id", "name", "status", "scenario_id", "action_id", "risk_id",
                 "likelihood_level", "business_impact", "data_impact", "compliance_impact",
                 "notes", "mitigation_pre", "mitigation_pro"]
    write_csv(data_dir / "risk_statement.csv", unique_rs, rs_fields)

    write_csv(data_dir / "rs_under_scenario.csv",
              [{"rs_id": r["rs_id"], "scenario_id": r["scenario_id"]} for r in unique_rs],
              ["rs_id", "scenario_id"])
    write_csv(data_dir / "rs_about_action.csv",
              [{"rs_id": r["rs_id"], "action_id": r["action_id"]} for r in unique_rs],
              ["rs_id", "action_id"])
    write_csv(data_dir / "rs_asserts_risk.csv",
              [{"rs_id": r["rs_id"], "risk_id": r["risk_id"]} for r in unique_rs],
              ["rs_id", "risk_id"])

    print(f"Wrote {len(scenarios)} scenarios, {len(action_options)} actions, {len(risks)} risks, {len(unique_rs)} statements to {data_dir}")


if __name__ == "__main__":
    main()
